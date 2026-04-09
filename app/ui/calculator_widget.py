"""Calcolatrice scientifica base (scheda Strumenti nel menu principale)."""

from __future__ import annotations

import math
import os
from datetime import date
from pathlib import Path
from typing import Callable

from PyQt6.QtCore import QEvent, QObject, Qt
from PyQt6.QtGui import QFont, QKeyEvent
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)


def _format_value(x: float) -> str:
    if math.isnan(x) or math.isinf(x):
        return "Errore"
    if abs(x) < 1e-15 and x != 0:
        return f"{x:.6e}"
    s = f"{x:.14g}"
    if "e" in s.lower():
        return s
    # Non usare rstrip("0") su interi tipo "100" (diventerebbe "1").
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    if s in {"", "-"}:
        s = "0"
    if s == "-0":
        s = "0"
    return s


class CalculatorWidget(QWidget):
    """Calcolatrice con operazioni base, memoria, radice, potenza, percentuale e tastiera."""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._accumulator: float | None = None
        self._pending_op: str | None = None
        self._current = "0"
        self._fresh = False
        self._memory = 0.0
        self._expr_line = ""
        self._build_ui()
        self._update_display()

    def _clear_error_state(self) -> None:
        if self._current != "Errore":
            return
        self._accumulator = None
        self._pending_op = None
        self._expr_line = ""
        self._current = "0"
        self._fresh = True

    def _parse_current(self) -> float:
        try:
            if self._current == "Errore":
                return 0.0
            if not self._current or self._current == "-":
                return 0.0
            return float(self._current.replace(",", "."))
        except ValueError:
            return 0.0

    def _set_current_from_float(self, x: float) -> None:
        self._current = _format_value(x)

    @staticmethod
    def _apply_binary(a: float, op: str, b: float) -> float:
        if op == "+":
            return a + b
        if op == "−":
            return a - b
        if op == "×":
            return a * b
        if op == "÷":
            if b == 0:
                raise ZeroDivisionError
            return a / b
        raise ValueError(op)

    def _commit_pending(self) -> None:
        if self._pending_op is None or self._accumulator is None:
            return
        b = self._parse_current()
        self._accumulator = self._apply_binary(self._accumulator, self._pending_op, b)
        self._set_current_from_float(self._accumulator)

    def _update_display(self) -> None:
        self.display.setText(self._current)
        self.expr_label.setText(self._expr_line)
        self.mem_label.setText("M" if abs(self._memory) > 1e-15 else "")

    def _digit(self, d: str) -> None:
        self._clear_error_state()
        if self._fresh:
            self._current = d
            self._fresh = False
        else:
            if self._current == "0" and d != ".":
                self._current = d
            elif self._current == "-0" and d != ".":
                self._current = "-" + d
            else:
                self._current += d
        self._expr_line = ""
        self._update_display()

    def _decimal(self) -> None:
        self._clear_error_state()
        if self._fresh:
            self._current = "0."
            self._fresh = False
        elif "." not in self._current and "," not in self._current:
            self._current += "."
        self._update_display()

    def _negate(self) -> None:
        self._clear_error_state()
        if self._fresh:
            self._current = "-0"
            self._fresh = False
        elif self._current.startswith("-"):
            self._current = self._current[1:] or "0"
        else:
            self._current = "-" + self._current if self._current != "0" else "-0"
        self._update_display()

    def _percent(self) -> None:
        self._clear_error_state()
        v = self._parse_current() / 100.0
        self._set_current_from_float(v)
        self._expr_line = ""
        self._update_display()

    def _sqrt(self) -> None:
        self._clear_error_state()
        v = self._parse_current()
        if v < 0:
            self._current = "Errore"
            self._expr_line = ""
            self._update_display()
            return
        self._set_current_from_float(math.sqrt(v))
        self._fresh = True
        self._expr_line = f"√({ _format_value(v) })"
        self._update_display()

    def _square(self) -> None:
        self._clear_error_state()
        v = self._parse_current()
        self._set_current_from_float(v * v)
        self._fresh = True
        self._expr_line = f"({ _format_value(v) })²"
        self._update_display()

    def _reciprocal(self) -> None:
        self._clear_error_state()
        v = self._parse_current()
        if v == 0:
            self._current = "Errore"
            self._expr_line = ""
            self._update_display()
            return
        self._set_current_from_float(1.0 / v)
        self._fresh = True
        self._expr_line = f"1/({ _format_value(v) })"
        self._update_display()

    def _binary_op(self, op: str) -> None:
        self._clear_error_state()
        cur = self._parse_current()
        if self._accumulator is None:
            self._accumulator = cur
        elif self._pending_op is not None and not self._fresh:
            try:
                self._accumulator = self._apply_binary(self._accumulator, self._pending_op, cur)
                self._set_current_from_float(self._accumulator)
            except ZeroDivisionError:
                self._current = "Errore"
                self._accumulator = None
                self._pending_op = None
                self._expr_line = ""
                self._update_display()
                return
        else:
            self._accumulator = cur
        self._pending_op = op
        self._expr_line = f"{_format_value(self._accumulator)} {op}"
        self._fresh = True
        self._update_display()

    def _equals(self) -> None:
        if self._current == "Errore":
            return
        if self._pending_op is None:
            return
        if self._accumulator is None:
            return
        try:
            b = self._parse_current()
            r = self._apply_binary(self._accumulator, self._pending_op, b)
            self._expr_line = f"{_format_value(self._accumulator)} {self._pending_op} {_format_value(b)} ="
            self._accumulator = None
            self._pending_op = None
            self._set_current_from_float(r)
            self._fresh = True
        except ZeroDivisionError:
            self._current = "Errore"
            self._accumulator = None
            self._pending_op = None
            self._expr_line = ""
        self._update_display()

    def _clear(self) -> None:
        self._accumulator = None
        self._pending_op = None
        self._current = "0"
        self._fresh = False
        self._expr_line = ""
        self._update_display()

    def _clear_entry(self) -> None:
        self._current = "0"
        self._fresh = False
        self._update_display()

    def _backspace(self) -> None:
        if self._current == "Errore":
            self._clear()
            return
        if self._fresh:
            return
        if len(self._current) <= 1 or (self._current.startswith("-") and len(self._current) <= 2):
            self._current = "0"
        else:
            self._current = self._current[:-1]
        self._update_display()

    def _mem_add(self) -> None:
        self._clear_error_state()
        self._memory += self._parse_current()

    def _mem_sub(self) -> None:
        self._clear_error_state()
        self._memory -= self._parse_current()

    def _mem_recall(self) -> None:
        self._clear_error_state()
        self._set_current_from_float(self._memory)
        self._fresh = False
        self._update_display()

    def _mem_clear(self) -> None:
        self._memory = 0.0
        self._update_display()

    def load_value_from_external(self, value: float) -> None:
        """Imposta il display come numero digitato (es. da tabella); resetta la catena di operazioni."""
        self._clear_error_state()
        self._accumulator = None
        self._pending_op = None
        self._expr_line = ""
        self._set_current_from_float(value)
        self._fresh = True
        self._update_display()

    def display_numeric_or_none(self) -> float | None:
        if self._current == "Errore":
            return None
        try:
            return float(self._current.replace(",", "."))
        except ValueError:
            return None

    def has_pending_binary_for_table(self) -> bool:
        """Dopo un operatore (+ − × ÷), il secondo operando può arrivare da una cella."""
        return self._pending_op is not None and self._accumulator is not None

    def apply_table_cell_as_second_operand(self, value: float) -> None:
        """Completa l'operazione in sospeso usando il valore della seconda cella."""
        self._clear_error_state()
        if self._pending_op is None or self._accumulator is None:
            self.load_value_from_external(value)
            return
        try:
            r = self._apply_binary(self._accumulator, self._pending_op, value)
        except ZeroDivisionError:
            self._current = "Errore"
            self._accumulator = None
            self._pending_op = None
            self._expr_line = ""
            self._update_display()
            return
        a_fmt = _format_value(self._accumulator)
        b_fmt = _format_value(value)
        self._expr_line = f"{a_fmt} {self._pending_op} {b_fmt} ="
        self._accumulator = None
        self._pending_op = None
        self._set_current_from_float(r)
        self._fresh = True
        self._update_display()

    def _make_btn(
        self,
        text: str,
        slot: Callable[[], None],
        *,
        primary: bool = False,
        op: bool = False,
    ) -> QPushButton:
        btn = QPushButton(text)
        btn.setMinimumHeight(38)
        btn.setMinimumWidth(52)
        f = QFont("Segoe UI", 10)
        if primary:
            f.setWeight(QFont.Weight.DemiBold)
        btn.setFont(f)
        if primary:
            btn.setObjectName("calcBtnPrimary")
        elif op:
            btn.setObjectName("calcBtnOp")
        else:
            btn.setObjectName("calcBtn")
        btn.clicked.connect(slot)
        return btn

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(10, 10, 10, 10)
        root.setSpacing(8)

        title = QLabel("Calcolatrice")
        title.setObjectName("sectionTitle")
        tf = QFont("Segoe UI", 12, QFont.Weight.Bold)
        title.setFont(tf)
        root.addWidget(title)

        top = QHBoxLayout()
        self.mem_label = QLabel("")
        self.mem_label.setFixedWidth(28)
        self.mem_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.mem_label.setObjectName("calcMem")
        top.addWidget(self.mem_label)
        self.expr_label = QLabel("")
        self.expr_label.setObjectName("calcExpr")
        self.expr_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        top.addWidget(self.expr_label, 1)
        root.addLayout(top)

        self.display = QLineEdit("0")
        self.display.setObjectName("calcDisplay")
        self.display.setReadOnly(True)
        self.display.setAlignment(Qt.AlignmentFlag.AlignRight)
        df = QFont("Consolas", 18)
        self.display.setFont(df)
        self.display.setMinimumHeight(48)
        root.addWidget(self.display)

        grid = QGridLayout()
        grid.setSpacing(6)

        r = 0
        grid.addWidget(self._make_btn("MC", self._mem_clear, op=True), r, 0)
        grid.addWidget(self._make_btn("M+", self._mem_add, op=True), r, 1)
        grid.addWidget(self._make_btn("M−", self._mem_sub, op=True), r, 2)
        grid.addWidget(self._make_btn("MR", self._mem_recall, op=True), r, 3)

        r = 1
        grid.addWidget(self._make_btn("%", self._percent, op=True), r, 0)
        grid.addWidget(self._make_btn("CE", self._clear_entry, op=True), r, 1)
        grid.addWidget(self._make_btn("C", self._clear, op=True), r, 2)
        grid.addWidget(self._make_btn("⌫", self._backspace, op=True), r, 3)

        r = 2
        grid.addWidget(self._make_btn("1/x", self._reciprocal, op=True), r, 0)
        grid.addWidget(self._make_btn("x²", self._square, op=True), r, 1)
        grid.addWidget(self._make_btn("√", self._sqrt, op=True), r, 2)
        grid.addWidget(self._make_btn("÷", lambda: self._binary_op("÷"), op=True), r, 3)

        r = 3
        for i, d in enumerate(["7", "8", "9"]):
            grid.addWidget(self._make_btn(d, lambda checked=False, x=d: self._digit(x)), r, i)
        grid.addWidget(self._make_btn("×", lambda: self._binary_op("×"), op=True), r, 3)

        r = 4
        for i, d in enumerate(["4", "5", "6"]):
            grid.addWidget(self._make_btn(d, lambda checked=False, x=d: self._digit(x)), r, i)
        grid.addWidget(self._make_btn("−", lambda: self._binary_op("−"), op=True), r, 3)

        r = 5
        for i, d in enumerate(["1", "2", "3"]):
            grid.addWidget(self._make_btn(d, lambda checked=False, x=d: self._digit(x)), r, i)
        grid.addWidget(self._make_btn("+", lambda: self._binary_op("+"), op=True), r, 3)

        r = 6
        grid.addWidget(self._make_btn("±", self._negate, op=True), r, 0)
        grid.addWidget(self._make_btn("0", lambda: self._digit("0")), r, 1)
        grid.addWidget(self._make_btn(",", self._decimal), r, 2)
        grid.addWidget(self._make_btn("=", self._equals, primary=True), r, 3)

        root.addLayout(grid)
        root.addStretch(0)

        self.setStyleSheet(
            """
            #calcDisplay {
                background: #0f172a;
                color: #f8fafc;
                border: 2px solid #334155;
                border-radius: 10px;
                padding: 8px 14px;
            }
            #calcExpr {
                color: #64748b;
                font-size: 13px;
                min-height: 20px;
            }
            #calcMem {
                color: #0f766e;
                font-weight: 800;
                font-size: 14px;
            }
            QPushButton#calcBtn {
                background: #ffffff;
                color: #0f172a;
                border: 1px solid #cbd5e1;
                border-radius: 10px;
            }
            QPushButton#calcBtn:hover {
                background: #f1f5f9;
                border-color: #94a3b8;
            }
            QPushButton#calcBtnOp {
                background: #e0f2fe;
                color: #0c4a6e;
                border: 1px solid #7dd3fc;
                border-radius: 10px;
                font-weight: 600;
            }
            QPushButton#calcBtnOp:hover {
                background: #bae6fd;
            }
            QPushButton#calcBtnPrimary {
                background: #0f766e;
                color: #ffffff;
                border: 1px solid #0f766e;
                border-radius: 10px;
            }
            QPushButton#calcBtnPrimary:hover {
                background: #0b5f58;
            }
            """
        )
        self.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.setMaximumWidth(318)
        self.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Preferred)

    def keyPressEvent(self, event: QKeyEvent) -> None:
        key = event.key()
        text = event.text()
        if key == Qt.Key.Key_Escape:
            self._clear()
            event.accept()
            return
        if key == Qt.Key.Key_Backspace:
            self._backspace()
            event.accept()
            return
        if key in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            self._equals()
            event.accept()
            return
        if Qt.Key.Key_0 <= key <= Qt.Key.Key_9:
            self._digit(str(key - int(Qt.Key.Key_0)))
            event.accept()
            return
        if text and text in "0123456789":
            self._digit(text)
            event.accept()
            return
        if text in ".,;":
            self._decimal()
            event.accept()
            return
        if text == "+":
            self._binary_op("+")
            event.accept()
            return
        if text == "-":
            self._binary_op("−")
            event.accept()
            return
        if text == "*":
            self._binary_op("×")
            event.accept()
            return
        if text == "/":
            self._binary_op("÷")
            event.accept()
            return
        if text == "%":
            self._percent()
            event.accept()
            return
        super().keyPressEvent(event)


def _parse_table_cell_float(item: QTableWidgetItem | None) -> float | None:
    if item is None:
        return None
    t = item.text().strip().replace(",", ".")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


class CalculatorWorkspaceWidget(QWidget):
    """Calcolatrice compatta a sinistra, tabella editabile a destra con collegamenti valore ↔ display."""

    def __init__(self, main_window: QWidget, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._main_window = main_window
        self._ctl = main_window.repository
        self._build_ui()
        self.value_table.installEventFilter(self)
        self._refresh_vista_combo()

    def _exports_dir(self) -> Path:
        settings = getattr(self._ctl, "settings", None)
        if settings is not None:
            return settings.get_calculator_export_base_path()
        local = os.getenv("LOCALAPPDATA") or os.path.expanduser("~")
        p = Path(local) / "HDManagerDesktop" / "calculator_exports"
        p.mkdir(parents=True, exist_ok=True)
        return p

    def _ensure_calcolatrice_tag(self) -> None:
        self._ctl.repository.upsert_tag(None, "Calcolatrice", "#0f766e", "")

    def _build_ui(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(16, 16, 16, 16)
        outer.setSpacing(0)

        card = QFrame()
        card.setObjectName("clientDashboardCard")
        card.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        card_l = QVBoxLayout(card)
        card_l.setContentsMargins(0, 0, 0, 0)
        card_l.setSpacing(0)

        header = QFrame()
        header.setObjectName("clientInfoCardHeader")
        header.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        hl = QHBoxLayout(header)
        hl.setContentsMargins(14, 10, 14, 10)
        hl.setSpacing(10)
        ic = QLabel("🔢")
        ic.setObjectName("clientInfoCardHeaderIcon")
        tl = QLabel("Calcolatrice e tabella valori")
        tl.setObjectName("clientInfoCardHeaderTitle")
        hl.addWidget(ic)
        hl.addWidget(tl, 1)
        card_l.addWidget(header)

        body = QWidget()
        body_l = QHBoxLayout(body)
        body_l.setContentsMargins(14, 14, 14, 14)
        body_l.setSpacing(14)

        self.calculator = CalculatorWidget()
        body_l.addWidget(self.calculator, 0, Qt.AlignmentFlag.AlignTop)

        right = QWidget()
        right_l = QVBoxLayout(right)
        right_l.setContentsMargins(0, 0, 0, 0)
        right_l.setSpacing(8)

        title = QLabel("Tabella valori")
        title.setObjectName("subSectionTitle")
        right_l.addWidget(title)

        hint = QLabel(
            "Prima cella: carica il numero nel display. Poi un operatore sulla calcolatrice (+, −, ×, ÷). "
            "Seconda cella: completa il calcolo con il valore della cella. "
            "Scrivi direttamente nella cella per modificarla. Frecce per spostarti. "
            "Canc: una cella elimina la riga; più celle svuotano le celle. "
            "«Salva vista» chiede solo il nome: il file Excel va nella cartella impostata in "
            "Impostazioni → Setup Strumenti e viene registrato in Archivio (tag Calcolatrice)."
        )
        hint.setObjectName("subText")
        hint.setWordWrap(True)
        right_l.addWidget(hint)

        vista_row = QHBoxLayout()
        vista_row.setSpacing(8)
        vista_lbl = QLabel("Viste salvate:")
        vista_lbl.setObjectName("subText")
        vista_row.addWidget(vista_lbl)
        self.views_combo = QComboBox()
        self.views_combo.setObjectName("archiveFilter")
        self.views_combo.setMinimumWidth(220)
        self.views_combo.currentIndexChanged.connect(self._on_vista_combo_changed)
        vista_row.addWidget(self.views_combo, 1)
        right_l.addLayout(vista_row)

        table_wrap = QFrame()
        table_wrap.setObjectName("clientDashboardTableWrap")
        table_wrap.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        tw = QVBoxLayout(table_wrap)
        tw.setContentsMargins(0, 0, 0, 0)

        self.value_table = QTableWidget(10, 5)
        self.value_table.setObjectName("calcValueTable")
        self.value_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.value_table.verticalHeader().setDefaultSectionSize(28)
        for c in range(5):
            self.value_table.setHorizontalHeaderItem(c, QTableWidgetItem(f"C{c + 1}"))
        self.value_table.setAlternatingRowColors(True)
        self.value_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.value_table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.value_table.setEditTriggers(
            QAbstractItemView.EditTrigger.AnyKeyPressed
            | QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
        )
        self.value_table.cellClicked.connect(self._on_value_cell_clicked)
        tw.addWidget(self.value_table, 1)
        right_l.addWidget(table_wrap, 1)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)

        save_btn = QPushButton("Salva valore")
        save_btn.setObjectName("primaryActionButton")
        save_btn.clicked.connect(self._save_display_to_cell)
        btn_row.addWidget(save_btn)

        save_vista_btn = QPushButton("Salva vista")
        save_vista_btn.setObjectName("primaryActionButton")
        save_vista_btn.setToolTip(
            "Salva nella Directory Salvataggi (Impostazioni → Setup Strumenti) e aggiunge il file in Archivio "
            "con tag Calcolatrice."
        )
        save_vista_btn.clicked.connect(self._save_vista_to_archive)
        btn_row.addWidget(save_vista_btn)

        btn_row.addStretch(1)
        right_l.addLayout(btn_row)

        body_l.addWidget(right, 1)
        card_l.addWidget(body, 1)
        outer.addWidget(card, 1)

        self.setStyleSheet(
            """
            #calcValueTable {
                background: #ffffff;
                border: none;
                border-radius: 8px;
                gridline-color: #e7edf3;
            }
            """
        )

    def eventFilter(self, obj: QObject, event: QEvent) -> bool:
        if obj is self.value_table and event.type() == QEvent.Type.KeyPress:
            ke = event
            if ke.key() == Qt.Key.Key_Delete:
                if self.value_table.state() == QAbstractItemView.State.EditingState:
                    return False
                self._handle_table_delete()
                return True
        return super().eventFilter(obj, event)

    def _handle_table_delete(self) -> None:
        idxs = self.value_table.selectedIndexes()
        if not idxs:
            return
        if len(idxs) == 1:
            self.value_table.removeRow(idxs[0].row())
            return
        seen: set[tuple[int, int]] = set()
        for ix in idxs:
            key = (ix.row(), ix.column())
            if key in seen:
                continue
            seen.add(key)
            it = self.value_table.item(ix.row(), ix.column())
            if it is None:
                it = QTableWidgetItem("")
                self.value_table.setItem(ix.row(), ix.column(), it)
            it.setText("")

    def _write_table_xlsx(self, path: Path) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Calcolatrice"
        for r in range(self.value_table.rowCount()):
            for c in range(self.value_table.columnCount()):
                item = self.value_table.item(r, c)
                val = item.text() if item is not None else ""
                ws.cell(row=r + 1, column=c + 1, value=val)
        wb.save(str(path))

    def _load_table_xlsx(self, path: str) -> None:
        from app.excel_io import load_workbook_openpyxl

        wb = load_workbook_openpyxl(path, read_only=True, data_only=True)
        ws = wb.active
        max_r = self.value_table.rowCount()
        max_c = self.value_table.columnCount()
        for r in range(max_r):
            for c in range(max_c):
                raw = ws.cell(row=r + 1, column=c + 1).value
                text = "" if raw is None else str(raw).strip()
                it = self.value_table.item(r, c)
                if it is None:
                    it = QTableWidgetItem(text)
                    self.value_table.setItem(r, c, it)
                else:
                    it.setText(text)
        wb.close()

    def _refresh_vista_combo(self) -> None:
        self.views_combo.blockSignals(True)
        self.views_combo.clear()
        self.views_combo.addItem("-- Seleziona una vista --", None)
        try:
            rows = self._ctl.archive.list_files_all()
        except Exception:
            rows = []
        for row in rows:
            tag = (row.get("tag_name") or "").strip()
            ext = (row.get("extension") or "").strip().lower()
            if tag != "Calcolatrice" or ext != "xlsx":
                continue
            fp = str(row.get("path") or "").strip()
            if not fp:
                continue
            label = str(row.get("name") or Path(fp).name)
            self.views_combo.addItem(label, fp)
        self.views_combo.blockSignals(False)

    def _on_vista_combo_changed(self, index: int) -> None:
        if index <= 0:
            return
        path = self.views_combo.currentData()
        if not path or not isinstance(path, str):
            return
        if not os.path.isfile(path):
            QMessageBox.warning(
                self,
                "Vista",
                "File non trovato sul disco. Potrebbe essere stato spostato o eliminato.",
            )
            return
        try:
            self._load_table_xlsx(path)
        except Exception as exc:
            QMessageBox.warning(self, "Vista", f"Impossibile aprire il file:\n{exc}")

    def _save_vista_to_archive(self) -> None:
        nome, ok = QInputDialog.getText(self, "Salva vista", "Nome della vista:")
        if not ok:
            return
        nome = (nome or "").strip()
        if not nome:
            QMessageBox.warning(self, "Salva vista", "Inserisci un nome.")
            return
        out_dir = self._exports_dir()
        try:
            self._ensure_calcolatrice_tag()
            day = date.today().strftime("%Y-%m-%d")
            base = f"{nome}_Calcolatrice_{day}.xlsx"
            path = out_dir / base
            n = 1
            while path.exists():
                path = out_dir / f"{nome}_Calcolatrice_{day}_{n}.xlsx"
                n += 1
            self._write_table_xlsx(path)
            self._ctl.archive.add_file(None, str(path.resolve()), "Calcolatrice")
        except Exception as exc:
            QMessageBox.warning(self, "Salva vista", str(exc))
            return
        self._refresh_vista_combo()
        if hasattr(self._main_window, "refresh_views"):
            self._main_window.refresh_views()
        QMessageBox.information(
            self,
            "Salva vista",
            f"Vista salvata in Archivio con tag «Calcolatrice»:\n{path}",
        )

    def _ensure_item(self, row: int, col: int) -> QTableWidgetItem:
        it = self.value_table.item(row, col)
        if it is None:
            it = QTableWidgetItem("")
            self.value_table.setItem(row, col, it)
        return it

    def _on_value_cell_clicked(self, row: int, col: int) -> None:
        item = self.value_table.item(row, col)
        v = _parse_table_cell_float(item)
        if v is None:
            return
        if self.calculator.has_pending_binary_for_table():
            self.calculator.apply_table_cell_as_second_operand(v)
        else:
            self.calculator.load_value_from_external(v)

    def _save_display_to_cell(self) -> None:
        num = self.calculator.display_numeric_or_none()
        if num is None:
            QMessageBox.warning(
                self,
                "Salva valore",
                "Il display non contiene un numero valido.",
            )
            return
        r = self.value_table.currentRow()
        c = self.value_table.currentColumn()
        if r < 0 or c < 0:
            QMessageBox.warning(self, "Salva valore", "Seleziona una cella nella tabella.")
            return
        it = self._ensure_item(r, c)
        it.setText(_format_value(num))
