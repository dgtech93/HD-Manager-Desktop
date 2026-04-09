"""Lettura file Excel (formati comuni) per la vista Strumenti Excel."""

from __future__ import annotations

import warnings
from pathlib import Path
from typing import Any


def load_workbook_openpyxl(*args: Any, **kwargs: Any):
    """
    Equivalente a openpyxl.load_workbook.

    I file con convalida dati (Data Validation) non pienamente supportata da openpyxl
    generano un UserWarning durante la lettura delle celle, non solo all’apertura:
    vedi `_register_openpyxl_warning_filters` sotto.
    """
    from openpyxl import load_workbook

    return load_workbook(*args, **kwargs)


def _register_openpyxl_warning_filters() -> None:
    """Registrato all’import: copre load_workbook e iter_rows / lettura foglio."""
    warnings.filterwarnings(
        "ignore",
        message=r".*Data Validation extension is not supported.*",
        category=UserWarning,
    )


_register_openpyxl_warning_filters()

# Anteprima / dialog: evita di caricare milioni di righe in memoria
DEFAULT_MAX_ROWS = 50_001

_EXCEL_OPENPYXL = {".xlsx", ".xlsm", ".xltx", ".xltm"}
_EXCEL_XLRD = {".xls"}


def excel_extensions_filter() -> str:
    return (
        "File Excel (*.xlsx *.xlsm *.xls *.xltx *.xltm *.xlsb);;"
        "Excel 2007+ (*.xlsx *.xlsm);;"
        "Excel 97-2003 (*.xls);;"
        "Tutti i file (*.*)"
    )


def list_sheet_names(path: str | Path) -> list[str]:
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(str(path))
    ext = path.suffix.lower()
    if ext in _EXCEL_XLRD:
        import xlrd

        book = xlrd.open_workbook(str(path))
        return list(book.sheet_names())
    if ext == ".xlsb":
        return _list_sheet_names_xlsb(path)
    if ext in _EXCEL_OPENPYXL or ext == "":
        wb = load_workbook_openpyxl(filename=str(path), read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    raise ValueError(f"Estensione non supportata: {ext}")


def _list_sheet_names_xlsb(path: Path) -> list[str]:
    try:
        from pyxlsb import open_workbook
    except ImportError as exc:
        raise ValueError(
            "Per i file .xlsb installa il pacchetto opzionale: pip install pyxlsb"
        ) from exc
    with open_workbook(str(path)) as wb:
        return list(wb.sheets)


def read_sheet_rows(
    path: str | Path,
    sheet_name: str,
    *,
    max_rows: int = DEFAULT_MAX_ROWS,
) -> tuple[list[list[Any]], int | None]:
    """
    Restituisce (righe come liste di celle, max_righe_stimate_o_None).
    Ogni riga è una lista di valori (stessa lunghezza dopo normalizzazione).
    """
    path = Path(path)
    ext = path.suffix.lower()
    if ext in _EXCEL_XLRD:
        return _read_xls(path, sheet_name, max_rows)
    if ext == ".xlsb":
        return _read_xlsb(path, sheet_name, max_rows)
    if ext in _EXCEL_OPENPYXL or ext == "":
        return _read_openpyxl(path, sheet_name, max_rows)
    raise ValueError(f"Estensione non supportata: {ext}")


def _read_openpyxl(path: Path, sheet_name: str, max_rows: int) -> tuple[list[list[Any]], int | None]:
    wb = load_workbook_openpyxl(filename=str(path), read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Foglio non trovato: {sheet_name}")
        ws = wb[sheet_name]
        rows: list[list[Any]] = []
        total_guess: int | None = None
        try:
            if ws.max_row is not None:
                total_guess = int(ws.max_row)
        except Exception:
            total_guess = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append(list(row))
        return rows, total_guess
    finally:
        wb.close()


def _read_xls(path: Path, sheet_name: str, max_rows: int) -> tuple[list[list[Any]], int | None]:
    import xlrd

    book = xlrd.open_workbook(str(path))
    try:
        sh = book.sheet_by_name(sheet_name)
        nrows = min(sh.nrows, max_rows)
        rows: list[list[Any]] = []
        for rx in range(nrows):
            row = [sh.cell_value(rx, cx) for cx in range(sh.ncols)]
            rows.append(row)
        return rows, int(sh.nrows)
    finally:
        book.release_resources()


def _read_xlsb(path: Path, sheet_name: str, max_rows: int) -> tuple[list[list[Any]], int | None]:
    try:
        from pyxlsb import open_workbook
    except ImportError as exc:
        raise ValueError("Per i file .xlsb: pip install pyxlsb") from exc
    rows: list[list[Any]] = []
    with open_workbook(str(path)) as wb:
        with wb.get_sheet(sheet_name) as sheet:
            for i, row in enumerate(sheet.rows()):
                if i >= max_rows:
                    break
                rows.append([c.v for c in row])
    return rows, None
